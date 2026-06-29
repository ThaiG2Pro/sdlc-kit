#!/usr/bin/env python3
"""Generate a QA test-case workbook from a testcases.json file.

Usage:
    python3 gen_testcases_xlsx.py <input.json> <output.xlsx>

Input JSON: an array of objects (extra keys ignored, missing keys → ""):
    {"test_id","technique","priority","objective","steps","expected",
     "requirement","automation","status"}
`steps` may be a list (joined with newlines) or a string.

Behavior:
  - If `openpyxl` is importable → writes a real .xlsx with a styled header row and the
    Status column colour-coded (Pass=green, Fail=red, N/A=grey).
  - If `openpyxl` is NOT installed → falls back to writing a .csv NEXT TO the requested
    output (same stem, .csv) so the deliverable still exists; exits 0 with a notice.
Zero third-party deps required for the fallback path (stdlib csv only).
"""
import csv
import json
import os
import sys

COLUMNS = [
    ("test_id", "Test ID"),
    ("technique", "Technique"),
    ("priority", "Priority"),
    ("objective", "Objective"),
    ("steps", "Steps"),
    ("expected", "Expected"),
    ("requirement", "Requirement"),
    ("automation", "Automation"),
    ("status", "Status"),
]
_STATUS_FILL = {"pass": "C6EFCE", "fail": "FFC7CE", "n/a": "D9D9D9", "na": "D9D9D9"}


def _cell(row: dict, key: str) -> str:
    v = row.get(key, "")
    if isinstance(v, (list, tuple)):
        return "\n".join(str(x) for x in v)
    return "" if v is None else str(v)


def _load(path: str):
    with open(path, "r", encoding="utf-8") as fh:
        data = json.load(fh)
    if isinstance(data, dict):  # tolerate {"test_cases":[...]} or {"testcases":[...]}
        data = data.get("test_cases") or data.get("testcases") or []
    if not isinstance(data, list):
        raise ValueError("input JSON must be an array of test-case objects")
    return data


def _write_csv(rows, out_csv: str) -> None:
    with open(out_csv, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow([h for _, h in COLUMNS])
        for r in rows:
            w.writerow([_cell(r, k) for k, _ in COLUMNS])


def _write_xlsx(rows, out_xlsx: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    ws.append([h for _, h in COLUMNS])
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    status_idx = [k for k, _ in COLUMNS].index("status") + 1
    for r in rows:
        ws.append([_cell(r, k) for k, _ in COLUMNS])
        fill = _STATUS_FILL.get(_cell(r, "status").strip().lower())
        if fill:
            ws.cell(row=ws.max_row, column=status_idx).fill = PatternFill("solid", fgColor=fill)
    for col, (key, header) in enumerate(COLUMNS, start=1):
        width = max(len(header), *(len(_cell(r, key)[:60]) for r in rows)) if rows else len(header)
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = min(max(width + 2, 10), 60)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(out_xlsx)


def main(argv) -> int:
    if len(argv) != 3:
        sys.stderr.write("usage: gen_testcases_xlsx.py <input.json> <output.xlsx>\n")
        return 2
    in_json, out_xlsx = argv[1], argv[2]
    try:
        rows = _load(in_json)
    except Exception as e:  # noqa: BLE001 — surface a clear message, fail closed
        sys.stderr.write(f"ERROR: cannot read {in_json}: {e}\n")
        return 1
    if not rows:
        sys.stderr.write(f"ERROR: {in_json} has 0 test cases — gate needs ≥1 row.\n")
        return 1
    os.makedirs(os.path.dirname(os.path.abspath(out_xlsx)) or ".", exist_ok=True)
    try:
        _write_xlsx(rows, out_xlsx)
        print(f"✓ wrote {out_xlsx} ({len(rows)} test cases)")
        return 0
    except ImportError:
        out_csv = os.path.splitext(out_xlsx)[0] + ".csv"
        _write_csv(rows, out_csv)
        sys.stderr.write(
            f"! openpyxl not installed — wrote {out_csv} instead ({len(rows)} test cases).\n"
            f"  For a real .xlsx: pip install openpyxl, then re-run.\n"
        )
        print(f"✓ wrote {out_csv} (CSV fallback, {len(rows)} test cases)")
        return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
